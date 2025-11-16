#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Create 3MB Excel file with 15 sheets"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime, timedelta
import random

def create_3mb_excel():
    """Create a 3MB Excel file with 15 sheets"""

    wb = Workbook()
    wb.remove(wb.active)

    print("Creating 3MB Excel file with 15 sheets...")

    # Sheet 1: 요약 대시보드
    print("  [1/15] 요약 대시보드")
    ws1 = wb.create_sheet("요약 대시보드")
    ws1['A1'] = 'CS 데일리 리포트 - 2025년 11월'
    ws1['A1'].font = Font(size=16, bold=True)
    ws1.merge_cells('A1:F1')

    headers1 = ['항목', '값', '전주 대비', '목표', '달성률', '비고']
    for col, header in enumerate(headers1, 1):
        ws1.cell(row=2, column=col, value=header).font = Font(bold=True)

    metrics = [
        ['총 문의 건수', 1247, '+15%', 1200, '103.9%', '목표 초과'],
        ['평균 응답 시간', 8.2, '-12%', 10, '82.0%', '목표 달성'],
        ['고객 만족도', 4.2, '+5%', 4.0, '105.0%', '목표 초과'],
    ]
    for row, metric in enumerate(metrics, 3):
        for col, value in enumerate(metric, 1):
            ws1.cell(row=row, column=col, value=value)

    # Sheets 2-11: 대용량 데이터 시트 (각각 10000행)
    categories = ['로그인 오류', '결제 실패', '기능 문의', 'API 연동', '데이터 오류',
                  '성능 이슈', '보안 문제', '계정 관리', '보고서 생성', '알림 설정']
    agents = ['김민수', '이영희', '박철수', '정수진', '최동욱']
    statuses = ['해결완료', '처리중', '보류', '에스컬레이션']
    channels = ['이메일', '전화', '채팅', '웹폼', '모바일앱']

    details_templates = [
        '고객이 로그인 시도 중 오류 발생. 비밀번호 재설정 링크 전송하여 해결.',
        '결제 진행 중 카드 승인 실패. 결제 게이트웨이 상태 확인 후 재시도 안내.',
        '데이터 내보내기 기능 사용 방법 문의. 상세 가이드 문서 링크 공유.',
        'API 토큰 만료 관련 문의. 새 토큰 발급 절차 안내.',
        '대시보드 로딩 시간이 느림. 데이터 필터링 최적화 방법 안내.',
        '2단계 인증 설정 오류. 인증앱 재설정 절차 안내하여 해결.',
        '월간 리포트 생성 시 일부 데이터 누락. 데이터 동기화 후 재생성.',
        '알림 이메일 수신 안됨. 스팸 필터 확인 및 이메일 주소 업데이트.',
    ]

    for sheet_num in range(2, 12):  # Sheets 2-11
        sheet_name = f"상세데이터_{sheet_num-1}"
        print(f"  [{sheet_num}/15] {sheet_name} (10000 rows)")
        ws = wb.create_sheet(sheet_name)

        # Headers
        headers = ['날짜', '시간', '문의ID', '고객명', '문의유형', '담당자',
                   '상태', '응답시간', '해결시간', '만족도', '채널', '상세내용']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Data rows - 10000 per sheet
        start_date = datetime(2025, 11, 1)
        for i in range(10000):
            row_num = i + 2
            date = start_date + timedelta(days=i % 30)

            ws.cell(row=row_num, column=1, value=date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=2, value=f"{random.randint(9,18)}:{random.randint(0,59):02d}")
            ws.cell(row=row_num, column=3, value=f"CS-{date.strftime('%Y%m%d')}-{i:05d}")
            ws.cell(row=row_num, column=4, value=f"고객{random.randint(1000,9999)}")
            ws.cell(row=row_num, column=5, value=random.choice(categories))
            ws.cell(row=row_num, column=6, value=random.choice(agents))
            ws.cell(row=row_num, column=7, value=random.choice(statuses))
            ws.cell(row=row_num, column=8, value=random.randint(3,30))
            ws.cell(row=row_num, column=9, value=random.randint(10,120))
            ws.cell(row=row_num, column=10, value=round(random.uniform(3.0,5.0), 1))
            ws.cell(row=row_num, column=11, value=random.choice(channels))
            ws.cell(row=row_num, column=12, value=random.choice(details_templates))

            if i % 2000 == 0 and i > 0:
                print(f"      Progress: {i}/10000 rows")

    # Sheet 12: 문의 유형별 통계
    print("  [12/15] 문의 유형별 통계")
    ws12 = wb.create_sheet("문의유형별통계")
    ws12['A1'] = '문의 유형별 통계 분석'
    ws12['A1'].font = Font(size=14, bold=True)

    headers12 = ['문의 유형', '총 건수', '비율', '평균 응답시간', '만족도']
    for col, header in enumerate(headers12, 1):
        ws12.cell(row=2, column=col, value=header).font = Font(bold=True)

    type_stats = [
        ['로그인 오류', 287, '23.0%', 7.5, 4.3],
        ['결제 실패', 156, '12.5%', 9.2, 4.0],
        ['기능 문의', 198, '15.9%', 8.8, 4.2],
        ['API 연동', 134, '10.7%', 12.3, 3.9],
        ['데이터 오류', 89, '7.1%', 10.5, 3.8],
    ]
    for row, stat in enumerate(type_stats, 3):
        for col, value in enumerate(stat, 1):
            ws12.cell(row=row, column=col, value=value)

    # Sheet 13: 담당자별 성과
    print("  [13/15] 담당자별 성과")
    ws13 = wb.create_sheet("담당자별성과")
    ws13['A1'] = '담당자별 성과 분석'
    ws13['A1'].font = Font(size=14, bold=True)

    headers13 = ['담당자', '처리 건수', '평균 응답시간', '고객 만족도']
    for col, header in enumerate(headers13, 1):
        ws13.cell(row=2, column=col, value=header).font = Font(bold=True)

    agent_stats = [
        ['김민수', 187, 7.8, 4.3],
        ['이영희', 165, 8.2, 4.2],
        ['박철수', 154, 8.5, 4.1],
        ['정수진', 172, 7.9, 4.3],
        ['최동욱', 149, 9.1, 4.0],
    ]
    for row, stat in enumerate(agent_stats, 3):
        for col, value in enumerate(stat, 1):
            ws13.cell(row=row, column=col, value=value)

    # Sheet 14: 시간대별 분석
    print("  [14/15] 시간대별 분석")
    ws14 = wb.create_sheet("시간대별분석")
    ws14['A1'] = '시간대별 문의 분석'
    ws14['A1'].font = Font(size=14, bold=True)

    headers14 = ['시간대', '문의 건수', '비율', '평균 응답시간']
    for col, header in enumerate(headers14, 1):
        ws14.cell(row=2, column=col, value=header).font = Font(bold=True)

    time_stats = [
        ['09:00-10:00', 156, '12.5%', 7.2],
        ['10:00-11:00', 189, '15.2%', 8.1],
        ['11:00-12:00', 142, '11.4%', 8.5],
        ['14:00-15:00', 134, '10.7%', 8.3],
        ['15:00-16:00', 167, '13.4%', 8.0],
    ]
    for row, stat in enumerate(time_stats, 3):
        for col, value in enumerate(stat, 1):
            ws14.cell(row=row, column=col, value=value)

    # Sheet 15: 주요 이슈 및 개선 방안
    print("  [15/15] 주요 이슈 및 개선방안")
    ws15 = wb.create_sheet("주요이슈및개선방안")
    ws15['A1'] = '주요 이슈 및 개선 방안'
    ws15['A1'].font = Font(size=14, bold=True)

    headers15 = ['순위', '이슈', '발생 빈도', '해결 방안', '담당자']
    for col, header in enumerate(headers15, 1):
        ws15.cell(row=2, column=col, value=header).font = Font(bold=True)

    issues = [
        [1, '비밀번호 재설정 프로세스 복잡', '높음', '자동 발송 시스템 구축', '이영희'],
        [2, '결제 게이트웨이 타임아웃', '중간', '재시도 로직 추가', '박철수'],
        [3, 'API 문서 불충분', '높음', 'API 문서 재작성', '정수진'],
        [4, '대시보드 성능 저하', '중간', '쿼리 최적화', '최동욱'],
        [5, '2단계 인증 오류', '낮음', 'UI 개선', '김민수'],
    ]
    for row, issue in enumerate(issues, 3):
        for col, value in enumerate(issue, 1):
            ws15.cell(row=row, column=col, value=value)

    # Save file
    filename = 'C:/Users/ohjh/ragflow/scripts/CS_Daily_Report_3MB.xlsx'
    print("\nSaving file...")
    wb.save(filename)

    # Check file size
    import os
    file_size = os.path.getsize(filename)
    print(f"\nExcel file created successfully!")
    print(f"  Filename: {filename}")
    print(f"  File size: {file_size / (1024*1024):.2f} MB")
    print(f"  Total sheets: {len(wb.sheetnames)}")
    print(f"  Total rows: ~100,000+ (10 sheets × 10,000 rows each)")

    return filename

if __name__ == "__main__":
    create_3mb_excel()
