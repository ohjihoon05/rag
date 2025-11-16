#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Create complex 3MB Excel file for RAGFlow testing"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import random

def create_complex_cs_report():
    """Create a complex CS daily report Excel file (target: ~3MB)"""

    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    print("Creating complex Excel file...")

    # Sheet 1: 요약 대시보드
    print("  - Sheet 1: 요약 대시보드")
    ws1 = wb.create_sheet("요약 대시보드")

    # Header
    ws1['A1'] = 'CS 데일리 리포트 - 2025년 11월 종합 분석'
    ws1['A1'].font = Font(size=16, bold=True)
    ws1['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws1['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws1.merge_cells('A1:F1')

    ws1['A3'] = '작성일'
    ws1['B3'] = datetime.now().strftime('%Y-%m-%d')
    ws1['A4'] = '작성자'
    ws1['B4'] = '김민수 (CS팀장)'
    ws1['A5'] = '리포트 기간'
    ws1['B5'] = '2025-11-01 ~ 2025-11-16'

    # 주요 지표
    ws1['A7'] = '주요 지표 요약'
    ws1['A7'].font = Font(size=14, bold=True)
    ws1['A7'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ws1.merge_cells('A7:F7')

    headers = ['항목', '값', '전주 대비', '목표', '달성률', '비고']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=8, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")

    metrics = [
        ['총 문의 건수', 1247, '+15%', 1200, '103.9%', '목표 초과'],
        ['평균 응답 시간 (분)', 8.2, '-12%', 10, '82.0%', '목표 달성'],
        ['평균 해결 시간 (분)', 24.5, '-8%', 30, '81.7%', '목표 달성'],
        ['고객 만족도', 4.2, '+5%', 4.0, '105.0%', '목표 초과'],
        ['1차 해결률', '84%', '+3%', '80%', '105.0%', '목표 초과'],
        ['에스컬레이션 비율', '11%', '-2%', '15%', '73.3%', '목표 달성'],
    ]

    for row, metric in enumerate(metrics, 9):
        for col, value in enumerate(metric, 1):
            ws1.cell(row=row, column=col, value=value)

    # Sheet 2: 일별 상세 데이터 (많은 행)
    print("  - Sheet 2: 일별 상세 데이터 (generating ~50000 rows)")
    ws2 = wb.create_sheet("일별 상세 데이터")

    headers2 = ['날짜', '시간', '문의 ID', '고객명', '문의 유형', '카테고리', '우선순위',
                '담당자', '상태', '응답시간(분)', '해결시간(분)', '만족도', '채널', '상세내용']
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    # Generate 15000 rows of data for ~3MB size
    categories = ['로그인 오류', '결제 실패', '기능 문의', 'API 연동', '데이터 오류',
                  '성능 이슈', '보안 문제', '계정 관리', '보고서 생성', '알림 설정']
    priorities = ['긴급', '높음', '보통', '낮음']
    agents = ['김민수', '이영희', '박철수', '정수진', '최동욱', '강미래', '윤서준', '한지원']
    statuses = ['해결완료', '처리중', '보류', '에스컬레이션']
    channels = ['이메일', '전화', '채팅', '웹폼', '모바일앱']

    details_templates = [
        '고객이 로그인 시도 중 "잘못된 비밀번호" 오류 메시지 발생. 비밀번호 재설정 링크 전송하여 해결.',
        '결제 진행 중 카드 승인 실패. 결제 게이트웨이 상태 확인 후 재시도 안내.',
        '데이터 내보내기 기능 사용 방법 문의. 상세 가이드 문서 링크 공유 및 화면 공유 지원.',
        'API 토큰 만료 관련 문의. 새 토큰 발급 절차 안내 및 문서 업데이트 필요성 확인.',
        '대시보드 로딩 시간이 느림. 데이터 필터링 최적화 방법 안내.',
        '2단계 인증 설정 오류. 인증앱 재설정 절차 안내하여 해결.',
        '월간 리포트 생성 시 일부 데이터 누락. 데이터 동기화 상태 확인 후 재생성.',
        '알림 이메일 수신 안됨. 스팸 필터 확인 및 이메일 주소 업데이트.',
        '사용자 권한 변경 요청. 관리자 승인 후 권한 업데이트 완료.',
        '계정 잠금 해제 요청. 보안 질문 확인 후 계정 활성화.',
    ]

    start_date = datetime(2025, 11, 1)

    for i in range(50000):  # 50000 rows for ~3MB target
        row_num = i + 2
        date = start_date + timedelta(days=i % 16)
        time = f"{random.randint(9, 18):02d}:{random.randint(0, 59):02d}"

        ws2.cell(row=row_num, column=1, value=date.strftime('%Y-%m-%d'))
        ws2.cell(row=row_num, column=2, value=time)
        ws2.cell(row=row_num, column=3, value=f"CS-{date.strftime('%Y%m%d')}-{i:05d}")
        ws2.cell(row=row_num, column=4, value=f"고객{random.randint(1000, 9999)}")
        ws2.cell(row=row_num, column=5, value=random.choice(categories))
        ws2.cell(row=row_num, column=6, value=random.choice(['기술지원', '계정관리', '결제/환불', '기능문의', '버그신고']))
        ws2.cell(row=row_num, column=7, value=random.choice(priorities))
        ws2.cell(row=row_num, column=8, value=random.choice(agents))
        ws2.cell(row=row_num, column=9, value=random.choice(statuses))
        ws2.cell(row=row_num, column=10, value=random.randint(3, 30))
        ws2.cell(row=row_num, column=11, value=random.randint(10, 120))
        ws2.cell(row=row_num, column=12, value=round(random.uniform(3.0, 5.0), 1))
        ws2.cell(row=row_num, column=13, value=random.choice(channels))
        ws2.cell(row=row_num, column=14, value=random.choice(details_templates))

        if i % 1000 == 0:
            print(f"    Generated {i}/5000 rows...")

    print("    Generated 5000/5000 rows")

    # Sheet 3: 문의 유형별 통계
    print("  - Sheet 3: 문의 유형별 통계")
    ws3 = wb.create_sheet("문의 유형별 통계")

    ws3['A1'] = '문의 유형별 통계 분석'
    ws3['A1'].font = Font(size=14, bold=True)
    ws3.merge_cells('A1:H1')

    headers3 = ['문의 유형', '총 건수', '비율', '평균 응답시간', '평균 해결시간',
                '만족도', '1차 해결률', '주요 해결 방안']
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    type_stats = [
        ['로그인 오류', 287, '23.0%', 7.5, 18.2, 4.3, '89%', '비밀번호 재설정, 계정 잠금 해제'],
        ['결제 실패', 156, '12.5%', 9.2, 28.5, 4.0, '78%', '결제 재시도, 카드 정보 확인'],
        ['기능 문의', 198, '15.9%', 8.8, 22.1, 4.2, '85%', '가이드 문서 제공, 화면 공유'],
        ['API 연동', 134, '10.7%', 12.3, 35.6, 3.9, '72%', 'API 문서 안내, 샘플 코드 제공'],
        ['데이터 오류', 89, '7.1%', 10.5, 42.3, 3.8, '68%', '데이터 재동기화, 수동 수정'],
        ['성능 이슈', 76, '6.1%', 15.2, 55.8, 3.7, '65%', '캐시 정리, 필터 최적화'],
        ['보안 문제', 112, '9.0%', 6.8, 15.4, 4.4, '92%', '2단계 인증 설정, 비밀번호 정책'],
        ['계정 관리', 145, '11.6%', 7.2, 16.8, 4.3, '88%', '프로필 업데이트, 권한 변경'],
        ['보고서 생성', 32, '2.6%', 18.5, 48.2, 3.6, '62%', '템플릿 수정, 데이터 검증'],
        ['알림 설정', 18, '1.4%', 8.5, 20.3, 4.1, '83%', '이메일 설정 확인, 필터 조정'],
    ]

    for row, stat in enumerate(type_stats, 3):
        for col, value in enumerate(stat, 1):
            ws3.cell(row=row, column=col, value=value)

    # Sheet 4: 담당자별 성과
    print("  - Sheet 4: 담당자별 성과")
    ws4 = wb.create_sheet("담당자별 성과")

    ws4['A1'] = '담당자별 성과 분석'
    ws4['A1'].font = Font(size=14, bold=True)
    ws4.merge_cells('A1:G1')

    headers4 = ['담당자', '처리 건수', '평균 응답시간', '평균 해결시간',
                '고객 만족도', '1차 해결률', '에스컬레이션']
    for col, header in enumerate(headers4, 1):
        cell = ws4.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    agent_stats = [
        ['김민수', 187, 7.8, 23.2, 4.3, '87%', '10건'],
        ['이영희', 165, 8.2, 24.8, 4.2, '85%', '12건'],
        ['박철수', 154, 8.5, 25.1, 4.1, '84%', '14건'],
        ['정수진', 172, 7.9, 23.8, 4.3, '86%', '11건'],
        ['최동욱', 149, 9.1, 26.2, 4.0, '82%', '16건'],
        ['강미래', 168, 8.0, 24.2, 4.2, '85%', '13건'],
        ['윤서준', 143, 8.8, 25.5, 4.1, '83%', '15건'],
        ['한지원', 109, 9.5, 27.8, 3.9, '80%', '18건'],
    ]

    for row, stat in enumerate(agent_stats, 3):
        for col, value in enumerate(stat, 1):
            ws4.cell(row=row, column=col, value=value)

    # Sheet 5: 시간대별 분석
    print("  - Sheet 5: 시간대별 분석")
    ws5 = wb.create_sheet("시간대별 분석")

    ws5['A1'] = '시간대별 문의 분석'
    ws5['A1'].font = Font(size=14, bold=True)
    ws5.merge_cells('A1:E1')

    headers5 = ['시간대', '문의 건수', '비율', '평균 응답시간', '주요 문의 유형']
    for col, header in enumerate(headers5, 1):
        cell = ws5.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="C5E0B4", end_color="C5E0B4", fill_type="solid")

    time_stats = [
        ['09:00-10:00', 156, '12.5%', 7.2, '로그인 오류, 계정 관리'],
        ['10:00-11:00', 189, '15.2%', 8.1, '로그인 오류, 기능 문의'],
        ['11:00-12:00', 142, '11.4%', 8.5, '결제 실패, 기능 문의'],
        ['12:00-13:00', 78, '6.3%', 9.2, '기능 문의'],
        ['13:00-14:00', 98, '7.9%', 8.8, '기능 문의, API 연동'],
        ['14:00-15:00', 134, '10.7%', 8.3, 'API 연동, 데이터 오류'],
        ['15:00-16:00', 167, '13.4%', 8.0, '결제 실패, 기능 문의'],
        ['16:00-17:00', 145, '11.6%', 8.4, '보고서 생성, 데이터 오류'],
        ['17:00-18:00', 138, '11.1%', 8.9, '기능 문의, 계정 관리'],
    ]

    for row, stat in enumerate(time_stats, 3):
        for col, value in enumerate(stat, 1):
            ws5.cell(row=row, column=col, value=value)

    # Sheet 6: 주요 이슈 및 해결 방안
    print("  - Sheet 6: 주요 이슈 및 해결 방안")
    ws6 = wb.create_sheet("주요 이슈 및 해결 방안")

    ws6['A1'] = '주요 이슈 및 해결 방안'
    ws6['A1'].font = Font(size=14, bold=True)
    ws6.merge_cells('A1:F1')

    headers6 = ['순위', '이슈', '발생 빈도', '심각도', '해결 방안', '예방 조치']
    for col, header in enumerate(headers6, 1):
        cell = ws6.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    issues = [
        [1, '비밀번호 재설정 프로세스 복잡', '높음', '중간',
         '재설정 링크 자동 발송 시스템 구축', 'FAQ 업데이트, 프로세스 간소화'],
        [2, '결제 게이트웨이 타임아웃', '중간', '높음',
         '타임아웃 설정 조정, 재시도 로직 추가', '게이트웨이 모니터링 강화'],
        [3, 'API 문서 불충분', '높음', '중간',
         'API 문서 전면 재작성, 샘플 코드 추가', '주기적 문서 리뷰'],
        [4, '대시보드 성능 저하', '중간', '중간',
         '쿼리 최적화, 캐싱 전략 도입', '성능 모니터링 도구 도입'],
        [5, '2단계 인증 설정 오류', '낮음', '높음',
         '설정 UI 개선, 상세 가이드 제공', '베타 테스트 강화'],
    ]

    for row, issue in enumerate(issues, 3):
        for col, value in enumerate(issue, 1):
            ws6.cell(row=row, column=col, value=value)

    # Sheet 7-16: 월별 상세 데이터 (10개월)
    print("  - Sheets 7-16: 월별 상세 데이터 (10 sheets)")
    months = ['1월', '2월', '3월', '4월', '5월', '6월', '7월', '8월', '9월', '10월']

    for month_idx, month in enumerate(months, 1):
        ws = wb.create_sheet(f"{month} 상세분석")
        ws['A1'] = f'2025년 {month} CS 리포트'
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')

        headers = ['일자', '문의건수', '평균응답시간', '만족도']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)

        # 30일 데이터
        for day in range(1, 31):
            ws.cell(row=day+2, column=1, value=f"{month_idx:02d}-{day:02d}")
            ws.cell(row=day+2, column=2, value=random.randint(30, 100))
            ws.cell(row=day+2, column=3, value=round(random.uniform(5, 15), 1))
            ws.cell(row=day+2, column=4, value=round(random.uniform(3.5, 5.0), 1))

    # Sheet 17: 채널별 분석
    print("  - Sheet 17: 채널별 분석")
    ws17 = wb.create_sheet("채널별 분석")
    ws17['A1'] = '채널별 문의 분석'
    ws17['A1'].font = Font(size=14, bold=True)
    ws17.merge_cells('A1:F1')

    headers17 = ['채널', '문의건수', '비율', '평균응답시간', '만족도', '주요 문의']
    for col, header in enumerate(headers17, 1):
        cell = ws17.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)

    channel_data = [
        ['이메일', 456, '36.6%', 12.3, 4.1, '기술문의, 계정관리'],
        ['전화', 389, '31.2%', 5.2, 4.3, '긴급문의, 결제문제'],
        ['채팅', 245, '19.6%', 3.8, 4.4, '간단문의, 사용법'],
        ['웹폼', 112, '9.0%', 18.5, 3.9, '기능요청, 버그신고'],
        ['모바일앱', 45, '3.6%', 7.2, 4.2, '앱오류, 푸시알림'],
    ]

    for row, data in enumerate(channel_data, 3):
        for col, value in enumerate(data, 1):
            ws17.cell(row=row, column=col, value=value)

    # Sheet 18: 우선순위별 분석
    print("  - Sheet 18: 우선순위별 분석")
    ws18 = wb.create_sheet("우선순위별 분석")
    ws18['A1'] = '우선순위별 처리 현황'
    ws18['A1'].font = Font(size=14, bold=True)
    ws18.merge_cells('A1:E1')

    headers18 = ['우선순위', '건수', '평균처리시간', 'SLA 준수율', '에스컬레이션']
    for col, header in enumerate(headers18, 1):
        cell = ws18.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)

    priority_data = [
        ['긴급', 89, '15분', '98%', '2건'],
        ['높음', 234, '45분', '95%', '8건'],
        ['보통', 678, '2시간', '92%', '15건'],
        ['낮음', 246, '4시간', '88%', '5건'],
    ]

    for row, data in enumerate(priority_data, 3):
        for col, value in enumerate(data, 1):
            ws18.cell(row=row, column=col, value=value)

    # Sheet 19: 고객 세그먼트 분석
    print("  - Sheet 19: 고객 세그먼트 분석")
    ws19 = wb.create_sheet("고객 세그먼트 분석")
    ws19['A1'] = '고객 세그먼트별 분석'
    ws19['A1'].font = Font(size=14, bold=True)
    ws19.merge_cells('A1:F1')

    headers19 = ['세그먼트', '고객수', '문의건수', '평균건수/고객', '만족도', 'LTV']
    for col, header in enumerate(headers19, 1):
        cell = ws19.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)

    segment_data = [
        ['Enterprise', 45, 567, 12.6, 4.5, '₩5,200,000'],
        ['SMB', 234, 456, 1.9, 4.2, '₩850,000'],
        ['Startup', 456, 189, 0.4, 4.0, '₩120,000'],
        ['Individual', 1234, 35, 0.03, 3.8, '₩25,000'],
    ]

    for row, data in enumerate(segment_data, 3):
        for col, value in enumerate(data, 1):
            ws19.cell(row=row, column=col, value=value)

    # Sheet 20: 제품별 문의
    print("  - Sheet 20: 제품별 문의")
    ws20 = wb.create_sheet("제품별 문의")
    ws20['A1'] = '제품별 문의 분석'
    ws20['A1'].font = Font(size=14, bold=True)
    ws20.merge_cells('A1:E1')

    headers20 = ['제품', '문의건수', '비율', '주요문의', '만족도']
    for col, header in enumerate(headers20, 1):
        cell = ws20.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)

    product_data = [
        ['RAGFlow Core', 456, '36.6%', '설치, 설정, 모델통합', 4.2],
        ['RAGFlow Cloud', 345, '27.7%', '계정, 결제, 성능', 4.3],
        ['RAGFlow API', 234, '18.8%', 'API문서, 인증, 제한', 4.0],
        ['RAGFlow Mobile', 123, '9.9%', '앱오류, 동기화', 3.9],
        ['RAGFlow Enterprise', 89, '7.1%', '커스터마이징, 통합', 4.4],
    ]

    for row, data in enumerate(product_data, 3):
        for col, value in enumerate(data, 1):
            ws20.cell(row=row, column=col, value=value)

    # Sheets 21-25: 주차별 트렌드 (5 weeks)
    print("  - Sheets 21-25: 주차별 트렌드 (5 sheets)")
    for week in range(1, 6):
        ws = wb.create_sheet(f"Week {week} 트렌드")
        ws['A1'] = f'Week {week} 트렌드 분석'
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')

        headers = ['요일', '문의건수', '평균응답시간', '만족도']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)

        days = ['월요일', '화요일', '수요일', '목요일', '금요일', '토요일', '일요일']
        for day_idx, day in enumerate(days, 3):
            ws.cell(row=day_idx, column=1, value=day)
            ws.cell(row=day_idx, column=2, value=random.randint(40, 120))
            ws.cell(row=day_idx, column=3, value=round(random.uniform(5, 12), 1))
            ws.cell(row=day_idx, column=4, value=round(random.uniform(3.8, 4.5), 1))

    # Sheet 26: SLA 성과
    print("  - Sheet 26: SLA 성과")
    ws26 = wb.create_sheet("SLA 성과")
    ws26['A1'] = 'SLA 준수 성과'
    ws26['A1'].font = Font(size=14, bold=True)
    ws26.merge_cells('A1:F1')

    headers26 = ['SLA 항목', '목표', '실제', '달성률', '위반건수', '조치사항']
    for col, header in enumerate(headers26, 1):
        cell = ws26.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)

    sla_data = [
        ['긴급문의 응답시간', '15분', '12분', '120%', 2, '우선처리 프로세스 강화'],
        ['일반문의 응답시간', '2시간', '1.8시간', '111%', 5, '-'],
        ['긴급문의 해결시간', '4시간', '3.5시간', '114%', 3, '-'],
        ['일반문의 해결시간', '24시간', '22시간', '109%', 12, '복잡한 케이스 프로세스 개선'],
        ['고객만족도', '4.0', '4.2', '105%', 0, '-'],
    ]

    for row, data in enumerate(sla_data, 3):
        for col, value in enumerate(data, 1):
            ws26.cell(row=row, column=col, value=value)

    # Sheet 27: 개선 계획
    print("  - Sheet 27: 개선 계획")
    ws27 = wb.create_sheet("개선 계획")
    ws27['A1'] = '향후 개선 계획'
    ws27['A1'].font = Font(size=14, bold=True)
    ws27.merge_cells('A1:F1')

    headers27 = ['순위', '개선 항목', '현황', '목표', '기한', '담당자']
    for col, header in enumerate(headers27, 1):
        cell = ws27.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)

    improvement_data = [
        [1, '챗봇 도입으로 1차 응답 자동화', '수동 응답 100%', '자동 응답 40%', '2025-12', '이영희'],
        [2, 'FAQ 시스템 개편', '검색률 30%', '검색률 60%', '2025-11', '박철수'],
        [3, 'API 문서 전면 개편', '만족도 3.9', '만족도 4.3', '2025-12', '정수진'],
        [4, '성능 모니터링 도구 도입', '수동 모니터링', '실시간 모니터링', '2026-01', '최동욱'],
        [5, '고객 포털 개선', 'Self-service 15%', 'Self-service 35%', '2026-02', '강미래'],
    ]

    for row, data in enumerate(improvement_data, 3):
        for col, value in enumerate(data, 1):
            ws27.cell(row=row, column=col, value=value)

    # Sheet 28: 팀 교육 현황
    print("  - Sheet 28: 팀 교육 현황")
    ws28 = wb.create_sheet("팀 교육 현황")
    ws28['A1'] = '팀 교육 및 역량 개발'
    ws28['A1'].font = Font(size=14, bold=True)
    ws28.merge_cells('A1:E1')

    headers28 = ['교육 과정', '참여 인원', '완료율', '만족도', '다음 일정']
    for col, header in enumerate(headers28, 1):
        cell = ws28.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)

    training_data = [
        ['고객 응대 스킬', 8, '100%', 4.5, '2025-12-15'],
        ['기술 제품 심화', 8, '87.5%', 4.3, '2025-11-30'],
        ['API 연동 가이드', 6, '100%', 4.4, '2025-12-10'],
        ['클레임 처리 기법', 8, '75%', 4.2, '진행중'],
        ['시스템 장애 대응', 8, '100%', 4.6, '2026-01-05'],
    ]

    for row, data in enumerate(training_data, 3):
        for col, value in enumerate(data, 1):
            ws28.cell(row=row, column=col, value=value)

    # Save file
    filename = 'C:/Users/ohjh/ragflow/scripts/CS_Daily_Report_Complex.xlsx'
    wb.save(filename)

    # Check file size
    import os
    file_size = os.path.getsize(filename)
    print(f"\nExcel file created: {filename}")
    print(f"File size: {file_size / (1024*1024):.2f} MB")
    print(f"Total sheets: {len(wb.sheetnames)}")
    print(f"Sheet names: {', '.join(wb.sheetnames)}")

    return filename

if __name__ == "__main__":
    create_complex_cs_report()
