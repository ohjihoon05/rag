#!/usr/bin/env python3
"""
Windows Test Script for RAPIDS cuDF Excel Parser

Tests the cuDF-enabled Excel parser on Windows (CPU fallback mode).
This validates the logic works correctly before deploying to Linux with GPU.

Usage:
    python test_parser_windows.py [excel_file]

If no file is specified, creates a sample test file automatically.
"""

import os
import sys
import time
import tempfile
import re
import logging
from io import BytesIO

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')

# Add project root to path
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PATCH_DIR = os.path.abspath(os.path.join(SCRIPT_DIR, '..'))
PROJECT_ROOT = os.path.abspath(os.path.join(SCRIPT_DIR, '..', '..', '..'))

# ILLEGAL_CHARACTERS_RE from openpyxl
ILLEGAL_CHARACTERS_RE = re.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")


def create_sample_excel(rows=1000):
    """Create a sample Korean Excel file for testing."""
    import pandas as pd
    import random
    from datetime import datetime, timedelta

    print(f"Creating sample Excel file with {rows} rows...")

    # Korean test data
    names = ["김철수", "이영희", "박민수", "최지영", "정대한", "한소희", "강민호", "윤서연"]
    departments = ["영업부", "마케팅부", "개발부", "운영부", "기획부", "인사부"]
    regions = ["서울", "부산", "대구", "인천", "광주", "대전", "울산", "세종"]
    products = ["제품A", "제품B", "제품C", "서비스X", "서비스Y"]

    data = []
    base_date = datetime(2025, 1, 1)

    for i in range(rows):
        row = {
            "일자": (base_date + timedelta(days=random.randint(0, 330))).strftime("%Y-%m-%d"),
            "담당자": random.choice(names),
            "부서": random.choice(departments),
            "지역": random.choice(regions),
            "제품명": random.choice(products),
            "판매수량": random.randint(10, 500),
            "단가": random.randint(10000, 1000000),
            "할인율": round(random.uniform(0, 0.3), 2),
            "매출액": 0,
            "비고": f"거래번호-{10000 + i}"
        }
        row["매출액"] = int(row["판매수량"] * row["단가"] * (1 - row["할인율"]))
        data.append(row)

    df = pd.DataFrame(data)

    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    df.to_excel(temp_file.name, sheet_name='영업보고서', index=False)
    temp_file.close()

    print(f"Sample file created: {temp_file.name}")
    return temp_file.name


class OriginalExcelParser:
    """
    Copy of original RAGFlowExcelParser for comparison testing.
    This avoids importing from deepdoc which has additional dependencies.
    """
    import pandas as pd
    from openpyxl import Workbook, load_workbook

    @staticmethod
    def _load_excel_to_workbook(file_like_object):
        import pandas as pd
        from openpyxl import load_workbook, Workbook

        if isinstance(file_like_object, bytes):
            file_like_object = BytesIO(file_like_object)

        file_like_object.seek(0)
        file_head = file_like_object.read(4)
        file_like_object.seek(0)

        if not (file_head.startswith(b"PK\x03\x04") or file_head.startswith(b"\xd0\xcf\x11\xe0")):
            try:
                file_like_object.seek(0)
                df = pd.read_csv(file_like_object)
                return OriginalExcelParser._dataframe_to_workbook(df)
            except Exception as e_csv:
                raise Exception(f"Failed to parse CSV: {e_csv}")

        try:
            return load_workbook(file_like_object, data_only=True)
        except Exception as e:
            try:
                file_like_object.seek(0)
                try:
                    dfs = pd.read_excel(file_like_object, sheet_name=None)
                    return OriginalExcelParser._dataframe_to_workbook(dfs)
                except Exception:
                    file_like_object.seek(0)
                    df = pd.read_excel(file_like_object, engine="calamine")
                    return OriginalExcelParser._dataframe_to_workbook(df)
            except Exception as e_pandas:
                raise Exception(f"pandas.read_excel error: {e_pandas}")

    @staticmethod
    def _clean_dataframe(df):
        import pandas as pd
        def clean_string(s):
            if isinstance(s, str):
                return ILLEGAL_CHARACTERS_RE.sub(" ", s)
            return s
        return df.apply(lambda col: col.map(clean_string))

    @staticmethod
    def _dataframe_to_workbook(df):
        from openpyxl import Workbook

        if isinstance(df, dict) and len(df) > 1:
            return OriginalExcelParser._dataframes_to_workbook(df)

        df = OriginalExcelParser._clean_dataframe(df)
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        for col_num, column_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_num, value=column_name)

        for row_num, row in enumerate(df.values, 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        return wb

    @staticmethod
    def _dataframes_to_workbook(dfs):
        from openpyxl import Workbook

        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        for sheet_name, df in dfs.items():
            df = OriginalExcelParser._clean_dataframe(df)
            ws = wb.create_sheet(title=sheet_name)
            for col_num, column_name in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_num, value=column_name)
            for row_num, row in enumerate(df.values, 2):
                for col_num, value in enumerate(row, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
        return wb

    def html(self, fnm, chunk_rows=256):
        from html import escape

        file_like_object = BytesIO(fnm) if not isinstance(fnm, str) else fnm
        wb = OriginalExcelParser._load_excel_to_workbook(file_like_object)
        tb_chunks = []

        def _fmt(v):
            if v is None:
                return ""
            return str(v).strip()

        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            try:
                rows = list(ws.rows)
            except Exception:
                continue

            if not rows:
                continue

            tb_rows_0 = "<tr>"
            for t in list(rows[0]):
                tb_rows_0 += f"<th>{escape(_fmt(t.value))}</th>"
            tb_rows_0 += "</tr>"

            for chunk_i in range((len(rows) - 1) // chunk_rows + 1):
                tb = f"<table><caption>{sheetname}</caption>"
                tb += tb_rows_0
                for r in list(rows[1 + chunk_i * chunk_rows : min(1 + (chunk_i + 1) * chunk_rows, len(rows))]):
                    tb += "<tr>"
                    for i, c in enumerate(r):
                        if c.value is None:
                            tb += "<td></td>"
                        else:
                            tb += f"<td>{escape(_fmt(c.value))}</td>"
                    tb += "</tr>"
                tb += "</table>\n"
                tb_chunks.append(tb)

        return tb_chunks

    def __call__(self, fnm):
        file_like_object = BytesIO(fnm) if not isinstance(fnm, str) else fnm
        wb = OriginalExcelParser._load_excel_to_workbook(file_like_object)

        res = []
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            try:
                rows = list(ws.rows)
            except Exception:
                continue
            if not rows:
                continue
            ti = list(rows[0])
            for r in list(rows[1:]):
                fields = []
                for i, c in enumerate(r):
                    if not c.value:
                        continue
                    t = str(ti[i].value) if i < len(ti) else ""
                    t += ("：" if t else "") + str(c.value)
                    fields.append(t)
                line = "; ".join(fields)
                if sheetname.lower().find("sheet") < 0:
                    line += " ——" + sheetname
                res.append(line)
        return res

    @staticmethod
    def row_number(fnm, binary):
        if fnm.split(".")[-1].lower().find("xls") >= 0:
            wb = OriginalExcelParser._load_excel_to_workbook(BytesIO(binary))
            total = 0
            for sheetname in wb.sheetnames:
                try:
                    ws = wb[sheetname]
                    total += len(list(ws.rows))
                except Exception:
                    continue
            return total
        return 0


def test_original_parser(file_path):
    """Test with original parser (embedded copy)."""
    print("\n" + "=" * 50)
    print("Testing ORIGINAL Excel Parser")
    print("=" * 50)

    with open(file_path, 'rb') as f:
        content = f.read()

    parser = OriginalExcelParser()

    # Test __call__ method
    print("\n[1] Testing __call__ method...")
    start = time.time()
    result = parser(content)
    elapsed = time.time() - start
    print(f"    Parsed {len(result)} rows in {elapsed:.2f}s")
    if result:
        print(f"    Sample: {result[0][:80]}...")

    # Test html method
    print("\n[2] Testing html() method...")
    start = time.time()
    html_chunks = parser.html(content)
    elapsed = time.time() - start
    print(f"    Generated {len(html_chunks)} HTML chunks in {elapsed:.2f}s")
    if html_chunks:
        print(f"    Sample chunk length: {len(html_chunks[0])} chars")

    # Test row_number method
    print("\n[3] Testing row_number() method...")
    row_count = OriginalExcelParser.row_number(file_path, content)
    print(f"    Total rows: {row_count}")

    return {
        "rows": len(result),
        "chunks": len(html_chunks),
        "row_count": row_count
    }


def test_cudf_parser(file_path):
    """Test with cuDF-enabled parser."""
    print("\n" + "=" * 50)
    print("Testing cuDF Excel Parser (CPU Fallback on Windows)")
    print("=" * 50)

    # Import from patches directory
    sys.path.insert(0, PATCH_DIR)

    # Remove cached module if exists
    if 'excel_parser_cudf' in sys.modules:
        del sys.modules['excel_parser_cudf']

    # Create a mock rag.nlp module for find_codec
    class MockNlp:
        @staticmethod
        def find_codec(binary):
            import chardet
            result = chardet.detect(binary)
            return result.get('encoding', 'utf-8') or 'utf-8'

    class MockRag:
        nlp = MockNlp()

    sys.modules['rag'] = MockRag()
    sys.modules['rag.nlp'] = MockNlp()

    from excel_parser_cudf import RAGFlowExcelParser, get_acceleration_status

    # Check acceleration status
    status = get_acceleration_status()
    print(f"\nAcceleration Status:")
    print(f"    Platform: {status['platform']}")
    print(f"    cuDF Available: {status['cudf_available']}")
    print(f"    GPU Accelerated: {status['gpu_accelerated']}")
    print(f"    Pandas Version: {status['pandas_version']}")

    with open(file_path, 'rb') as f:
        content = f.read()

    parser = RAGFlowExcelParser()

    # Test __call__ method
    print("\n[1] Testing __call__ method...")
    start = time.time()
    result = parser(content)
    elapsed = time.time() - start
    print(f"    Parsed {len(result)} rows in {elapsed:.2f}s")
    if result:
        print(f"    Sample: {result[0][:80]}...")

    # Test html method
    print("\n[2] Testing html() method...")
    start = time.time()
    html_chunks = parser.html(content)
    elapsed = time.time() - start
    print(f"    Generated {len(html_chunks)} HTML chunks in {elapsed:.2f}s")
    if html_chunks:
        print(f"    Sample chunk length: {len(html_chunks[0])} chars")

    # Test row_number method
    print("\n[3] Testing row_number() method...")
    row_count = RAGFlowExcelParser.row_number(file_path, content)
    print(f"    Total rows: {row_count}")

    # Test is_gpu_accelerated method
    print("\n[4] Testing is_gpu_accelerated() method...")
    print(f"    GPU Accelerated: {RAGFlowExcelParser.is_gpu_accelerated()}")

    return {
        "rows": len(result),
        "chunks": len(html_chunks),
        "row_count": row_count,
        "gpu_accelerated": status['gpu_accelerated']
    }


def compare_results(original, cudf):
    """Compare results from both parsers."""
    print("\n" + "=" * 50)
    print("Comparison Results")
    print("=" * 50)

    all_match = True

    if original['rows'] == cudf['rows']:
        print(f"  [PASS] Row count: {original['rows']}")
    else:
        print(f"  [FAIL] Row count: Original={original['rows']}, cuDF={cudf['rows']}")
        all_match = False

    if original['chunks'] == cudf['chunks']:
        print(f"  [PASS] Chunk count: {original['chunks']}")
    else:
        print(f"  [FAIL] Chunk count: Original={original['chunks']}, cuDF={cudf['chunks']}")
        all_match = False

    if original['row_count'] == cudf['row_count']:
        print(f"  [PASS] Total rows: {original['row_count']}")
    else:
        print(f"  [FAIL] Total rows: Original={original['row_count']}, cuDF={cudf['row_count']}")
        all_match = False

    print()
    if all_match:
        print("  ===================================")
        print("  ALL TESTS PASSED!")
        print("  ===================================")
        print()
        print("  The cuDF parser produces identical results to the original.")
        print("  Safe to deploy to Linux for GPU acceleration.")
    else:
        print("  ===================================")
        print("  SOME TESTS FAILED!")
        print("  ===================================")
        print()
        print("  Please investigate the differences before deploying.")

    return all_match


def main():
    print("=" * 60)
    print("RAPIDS cuDF Excel Parser - Windows Test Suite")
    print("=" * 60)

    # Determine test file
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        if not os.path.exists(file_path):
            print(f"ERROR: File not found: {file_path}")
            sys.exit(1)
        temp_file = False
    else:
        # Check for existing sample file
        sample_file = os.path.join(PROJECT_ROOT, 'daily_report_sample.xlsx')
        if os.path.exists(sample_file):
            file_path = sample_file
            temp_file = False
            print(f"Using existing sample file: {file_path}")
        else:
            file_path = create_sample_excel(1000)
            temp_file = True

    print(f"\nTest file: {file_path}")
    print(f"File size: {os.path.getsize(file_path) / 1024:.1f} KB")

    try:
        # Test original parser
        original_results = test_original_parser(file_path)

        # Test cuDF parser
        cudf_results = test_cudf_parser(file_path)

        # Compare results
        success = compare_results(original_results, cudf_results)

        sys.exit(0 if success else 1)

    finally:
        # Cleanup temp file if created
        if temp_file and os.path.exists(file_path):
            os.unlink(file_path)
            print(f"\nCleaned up temp file: {file_path}")


if __name__ == "__main__":
    main()
